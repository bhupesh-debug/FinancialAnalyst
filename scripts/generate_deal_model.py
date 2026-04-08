"""
generate_deal_model.py
======================
Generates a professional CRE deal model Excel workbook using openpyxl.

Output: /home/user/workspace/FinancialAnalyst/output/sample_deal_model.xlsx

Sheets
------
  1. Summary      — Deal overview KPIs with navigation index
  2. Rent Roll    — Formatted rent roll with unit details and analytics
  3. Pro Forma    — 5-year income statement with live Excel formulas
  4. Debt Sizing  — 3-lender comparison with max loan and recommendation
  5. Sensitivity  — One-way and two-way sensitivity tables with conditional formatting

Color-Coding Convention (CRE industry standard)
------------------------------------------------
  Blue text   — Hardcoded inputs (editable by user)
  Black text  — Formulas (calculated, do not override)
  Green text  — Cross-sheet references
"""

from __future__ import annotations

import os
import sys
from datetime import date

# Ensure models/ is on the path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.cre_deal_model import CREDealModel, OperatingExpenses, MarketAssumptions
from models.debt_sizing_engine import DebtSizingEngine, DEFAULT_LENDERS
from models.sensitivity_analysis import SensitivityAnalysis

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as xl_numbers,
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


# ---------------------------------------------------------------------------
# Constants — Colors
# ---------------------------------------------------------------------------
INPUT_FONT_COLOR = "0000FF"    # Blue — hardcoded inputs
FORMULA_FONT_COLOR = "000000"  # Black — formula outputs
XSHEET_FONT_COLOR = "008000"   # Green — cross-sheet links

HEADER_BG = "1F3864"           # Dark navy — primary header
HEADER_FG = "FFFFFF"           # White text on header
SUBHEADER_BG = "D6DCE4"        # Light steel — sub-headers
SECTION_BG = "EBF3FB"          # Light blue — section highlights
TOTAL_BG = "E2EFDA"            # Light green — total rows
BREACH_BG = "FFCCCC"           # Light red — DSCR breach
WARNING_BG = "FFFACD"          # Light yellow — warning

BORDER_COLOR = "B8B8B8"

# ---------------------------------------------------------------------------
# Style Helpers
# ---------------------------------------------------------------------------

def header_font(bold=True, size=11, color=HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def input_font(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size, color=INPUT_FONT_COLOR)

def formula_font(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size, color=FORMULA_FONT_COLOR)

def xsheet_font(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size, color=XSHEET_FONT_COLOR)

def header_fill(color=HEADER_BG):
    return PatternFill("solid", fgColor=color)

def section_fill(color=SECTION_BG):
    return PatternFill("solid", fgColor=color)

def total_fill():
    return PatternFill("solid", fgColor=TOTAL_BG)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(indent=1, wrap=False):
    return Alignment(horizontal="left", vertical="center", indent=indent, wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")

def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(bottom=s)

def thick_bottom_border():
    s = Side(style="medium", color="000000")
    return Border(bottom=s)

def write_cell(ws, row, col, value, font=None, fill=None, alignment=None,
               number_format=None, border=None):
    """Write a value to a cell with full style options."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    return cell

def section_header(ws, row, col, text, span_end_col=None):
    """Write a section header row spanning optional columns."""
    cell = write_cell(
        ws, row, col, text,
        font=header_font(bold=True, size=10),
        fill=header_fill(),
        alignment=left_align(indent=1),
    )
    if span_end_col and span_end_col > col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=span_end_col,
        )
    ws.row_dimensions[row].height = 18


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Build the Deal Model
# ---------------------------------------------------------------------------

def build_deal_model():
    """Construct the sample 8-unit NYC mixed-use deal model."""
    RENT_ROLL = [
        {"unit_id": "R-001", "tenant": "Café Delancey LLC",       "unit_type": "Retail",  "sf": 1_200, "monthly_rent": 7_000,  "lease_start": "2022-02-01", "lease_end": "2027-01-31", "annual_escalation": 0.030, "status": "Occupied"},
        {"unit_id": "R-002", "tenant": "LES Boutique Inc.",        "unit_type": "Retail",  "sf": 800,   "monthly_rent": 4_667,  "lease_start": "2021-06-01", "lease_end": "2026-05-31", "annual_escalation": 0.025, "status": "Occupied"},
        {"unit_id": "R-003", "tenant": "Pharmacare Plus",          "unit_type": "Retail",  "sf": 1_500, "monthly_rent": 9_375,  "lease_start": "2023-03-01", "lease_end": "2033-02-28", "annual_escalation": 0.020, "status": "Occupied"},
        {"unit_id": "O-101", "tenant": "Fintech Partners LP",      "unit_type": "Office",  "sf": 2_500, "monthly_rent": 10_417, "lease_start": "2023-01-01", "lease_end": "2028-12-31", "annual_escalation": 0.030, "status": "Occupied"},
        {"unit_id": "O-102", "tenant": "Creative Studio NYC",      "unit_type": "Office",  "sf": 1_800, "monthly_rent": 7_200,  "lease_start": "2020-09-01", "lease_end": "2025-08-31", "annual_escalation": 0.020, "status": "Occupied"},
        {"unit_id": "O-103", "tenant": "Meridian Advisory Group",  "unit_type": "Office",  "sf": 2_200, "monthly_rent": 9_350,  "lease_start": "2022-07-01", "lease_end": "2027-06-30", "annual_escalation": 0.030, "status": "Occupied"},
        {"unit_id": "O-104", "tenant": "",                         "unit_type": "Office",  "sf": 1_600, "monthly_rent": 0,      "lease_start": "2024-01-01", "lease_end": "2024-01-01", "annual_escalation": 0.030, "status": "Vacant"},
        {"unit_id": "S-001", "tenant": "SelfStore Manhattan",      "unit_type": "Storage", "sf": 400,   "monthly_rent": 600,    "lease_start": "2023-06-01", "lease_end": "2026-05-31", "annual_escalation": 0.030, "status": "Occupied"},
    ]
    total_sf = sum(r["sf"] for r in RENT_ROLL)
    opex = OperatingExpenses(
        real_estate_taxes=95_000,
        insurance=18_000,
        repairs_maintenance=22_000,
        management_fee_pct=0.04,
        utilities=14_000,
        general_admin=12_000,
        reserves=total_sf * 0.50,
    )
    market = MarketAssumptions(
        market_rent_per_sf={"Retail": 68.0, "Office": 50.0, "Storage": 18.0},
        market_vacancy_rate=0.05,
        credit_loss_rate=0.01,
        lease_up_months=3,
    )
    model = CREDealModel(
        property_name="123 Delancey Street",
        address="123 Delancey St, New York, NY 10002",
        property_type="Mixed-Use (Retail + Office)",
        rent_roll=RENT_ROLL,
        opex=opex,
        market=market,
        analysis_date=date(2025, 1, 1),
    )
    return model


# ---------------------------------------------------------------------------
# Sheet 1 — Summary
# ---------------------------------------------------------------------------

def build_summary_sheet(ws, model: CREDealModel, sa: SensitivityAnalysis, base: dict):
    """Deal overview: key metrics, navigation links."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8

    # --- Title ---
    ws.merge_cells("B2:I2")
    write_cell(ws, 2, 2, "123 DELANCEY STREET — DEAL SUMMARY",
               font=Font(name="Calibri", bold=True, size=16, color=HEADER_FG),
               fill=header_fill("1F3864"),
               alignment=center_align())
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:I3")
    write_cell(ws, 3, 2, "123 Delancey St, New York, NY 10002  |  Mixed-Use (Retail + Office)  |  CRE Capital Markets Underwriting",
               font=Font(name="Calibri", size=9, italic=True, color="4F4F4F"),
               fill=PatternFill("solid", fgColor="EBF3FB"),
               alignment=center_align())
    ws.row_dimensions[3].height = 16

    # --- Navigation Index ---
    row = 5
    ws.merge_cells(f"B{row}:I{row}")
    write_cell(ws, row, 2, "WORKBOOK NAVIGATION",
               font=header_font(size=10), fill=header_fill(),
               alignment=left_align(indent=1))
    ws.row_dimensions[row].height = 18
    row += 1

    nav_items = [
        ("Rent Roll",    "Rent Roll",    "B",  "Full rent roll with unit-level detail"),
        ("Pro Forma",    "Pro Forma",    "C",  "5-year income statement pro forma"),
        ("Debt Sizing",  "Debt Sizing",  "D",  "3-lender comparison and recommendation"),
        ("Sensitivity",  "Sensitivity",  "E",  "One-way and two-way sensitivity analysis"),
    ]
    for sheet_name, display, col_hint, desc in nav_items:
        cell = ws.cell(row=row, column=2, value=f"→ {display}")
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1")
        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
        cell.alignment = left_align(indent=2)
        ws.merge_cells(f"C{row}:I{row}")
        write_cell(ws, row, 3, desc, font=formula_font(size=9),
                   alignment=left_align(), fill=PatternFill("solid", fgColor="F9F9F9"))
        ws.row_dimensions[row].height = 16
        row += 1

    # --- Key Assumptions Block ---
    row += 1
    section_header(ws, row, 2, "KEY UNDERWRITING ASSUMPTIONS", span_end_col=9)
    row += 1

    CAP_RATE = 0.055   # input
    DISC_RATE = 0.08
    TERM_CAP  = 0.060

    assumptions = [
        ("Analysis Date",          "2025-01-01",    None,       "Input",   None),
        ("Going-In Cap Rate",       CAP_RATE,        "0.0%",     "Input",   "Market cap rate for direct capitalization"),
        ("Discount Rate (DCF)",     DISC_RATE,       "0.0%",     "Input",   "Unlevered target IRR"),
        ("Terminal Cap Rate",        TERM_CAP,       "0.0%",     "Input",   "Exit cap rate at end of hold period"),
        ("OpEx Growth Rate",        0.03,            "0.0%",     "Input",   "Annual expense inflation assumption"),
        ("Market Vacancy Rate",     0.05,            "0.0%",     "Input",   "Stabilized vacancy allowance (market)"),
        ("Credit Loss Rate",        0.01,            "0.0%",     "Input",   "Bad debt allowance"),
    ]

    write_cell(ws, row, 2, "Assumption",   font=Font(name="Calibri", bold=True, size=9), fill=PatternFill("solid", fgColor=SUBHEADER_BG), alignment=center_align())
    write_cell(ws, row, 3, "Value",        font=Font(name="Calibri", bold=True, size=9), fill=PatternFill("solid", fgColor=SUBHEADER_BG), alignment=center_align())
    write_cell(ws, row, 4, "Type",         font=Font(name="Calibri", bold=True, size=9), fill=PatternFill("solid", fgColor=SUBHEADER_BG), alignment=center_align())
    write_cell(ws, row, 5, "Note",         font=Font(name="Calibri", bold=True, size=9), fill=PatternFill("solid", fgColor=SUBHEADER_BG), alignment=center_align())
    ws.merge_cells(f"E{row}:I{row}")
    ws.row_dimensions[row].height = 16
    row += 1

    for label, val, fmt, type_, note in assumptions:
        write_cell(ws, row, 2, label,     font=formula_font(size=9), alignment=left_align())
        cell = write_cell(ws, row, 3, val, font=input_font(size=9), alignment=right_align())
        if fmt:
            cell.number_format = fmt
        write_cell(ws, row, 4, type_,    font=formula_font(size=9), alignment=center_align())
        if note:
            ws.merge_cells(f"E{row}:I{row}")
            write_cell(ws, row, 5, note,  font=Font(name="Calibri", size=9, italic=True, color="666666"), alignment=left_align())
        ws.row_dimensions[row].height = 15
        row += 1

    # --- Key Metrics Block ---
    row += 1
    section_header(ws, row, 2, "KEY FINANCIAL METRICS", span_end_col=9)
    row += 1

    noi = model.net_operating_income()
    value = model.direct_cap_value(CAP_RATE)
    egi = model.effective_gross_income()
    gpr = model.gross_potential_rent()
    opex_total = model.total_operating_expenses()
    engine = DebtSizingEngine(noi=noi, property_value=value)
    rec = engine.recommend()

    metrics = [
        ("Total Rentable SF",           f"{model.total_sf:,.0f}",         None,       "sf"),
        ("Physical Occupancy",          f"{model.physical_occupancy:.1%}", None,       "Based on in-place leases"),
        ("Gross Potential Rent (GPR)",  f"${gpr:,.0f}",                   None,       "Annualized at 100% occupancy"),
        ("Effective Gross Income (EGI)",f"${egi:,.0f}",                   None,       "GPR less vacancy & credit loss"),
        ("Total Operating Expenses",    f"(${opex_total:,.0f})",           None,       "All property-level expenses"),
        ("Net Operating Income (NOI)",  f"${noi:,.0f}",                   None,       "EGI − OpEx"),
        ("NOI / SF",                    f"${model.noi_per_sf():.2f}/sf",   None,       ""),
        ("Direct Cap Value",            f"${value:,.0f}",                 None,       f"NOI / {CAP_RATE:.1%} cap rate"),
        ("Going-In Cap Rate",           f"{CAP_RATE:.2%}",                 None,       ""),
        ("Max Loan (Recommended)",      f"${rec['max_loan']:,.0f}",        None,       f"{rec['lender']} — {rec['binding_constraint']} binding"),
        ("Loan-to-Value (LTV)",         f"{rec['actual_ltv']:.1%}",        None,       ""),
        ("Debt Service Coverage (DSCR)",f"{rec['actual_dscr']:.2f}x",      None,       "NOI / Annual Debt Service"),
        ("Debt Yield",                  f"{rec['actual_debt_yield']:.1%}",  None,       "NOI / Loan Amount"),
        ("Annual Debt Service",         f"${rec['annual_debt_service']:,.0f}", None,   f"{rec['interest_rate']:.2%} rate"),
    ]

    # Column headers
    for col_idx, hdr in enumerate(["Metric", "Value", "Note"], start=2):
        bg = SUBHEADER_BG
        cell = write_cell(ws, row, col_idx, hdr,
                          font=Font(name="Calibri", bold=True, size=9),
                          fill=PatternFill("solid", fgColor=bg),
                          alignment=center_align())
    ws.merge_cells(f"D{row}:I{row}")
    ws.row_dimensions[row].height = 16
    row += 1

    for label, val, fmt, note in metrics:
        is_total = label in ("Net Operating Income (NOI)", "Direct Cap Value")
        bg_fill = total_fill() if is_total else None
        fnt = Font(name="Calibri", bold=is_total, size=9,
                   color=FORMULA_FONT_COLOR if not is_total else "000000")
        write_cell(ws, row, 2, label, font=fnt, fill=bg_fill, alignment=left_align())
        write_cell(ws, row, 3, val,   font=fnt, fill=bg_fill, alignment=right_align())
        if note:
            ws.merge_cells(f"D{row}:I{row}")
            write_cell(ws, row, 4, note,
                       font=Font(name="Calibri", size=9, italic=True, color="666666"),
                       fill=bg_fill, alignment=left_align())
        ws.row_dimensions[row].height = 15
        row += 1

    # --- Footer ---
    row += 1
    ws.merge_cells(f"B{row}:I{row}")
    write_cell(ws, row, 2,
               f"Source: Internal CRE Underwriting Model  |  Generated: {date.today().isoformat()}  |  For Institutional Use Only",
               font=Font(name="Calibri", size=8, italic=True, color="888888"),
               alignment=left_align())

    # Column widths
    set_col_width(ws, 2, 35)
    set_col_width(ws, 3, 18)
    set_col_width(ws, 4, 12)
    for c in range(5, 10):
        set_col_width(ws, c, 18)

    ws.freeze_panes = "B4"


# ---------------------------------------------------------------------------
# Sheet 2 — Rent Roll
# ---------------------------------------------------------------------------

def build_rent_roll_sheet(ws, model: CREDealModel):
    """Formatted rent roll table with live Excel formulas."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8

    # Title
    ws.merge_cells("B2:M2")
    write_cell(ws, 2, 2, "RENT ROLL — 123 DELANCEY STREET",
               font=Font(name="Calibri", bold=True, size=14, color=HEADER_FG),
               fill=header_fill(), alignment=center_align())
    ws.row_dimensions[2].height = 28

    # Sub-header
    ws.merge_cells("B3:M3")
    write_cell(ws, 3, 2, f"As of January 1, 2025  |  {model.total_sf:,.0f} SF Total  |  {model.physical_occupancy:.1%} Occupied",
               font=Font(name="Calibri", size=9, italic=True),
               fill=PatternFill("solid", fgColor=SECTION_BG),
               alignment=center_align())
    ws.row_dimensions[3].height = 14

    # Column headers (row 5)
    headers = [
        "Unit ID", "Tenant Name", "Type", "SF",
        "Mo. Rent ($)", "Ann. Rent ($)", "Rent/SF/Yr",
        "Lease Start", "Lease End", "Escalation",
        "Status", "Mos. Remaining", "% of GPR"
    ]
    hdr_row = 5
    for col_idx, h in enumerate(headers, start=2):
        write_cell(ws, hdr_row, col_idx, h,
                   font=Font(name="Calibri", bold=True, size=9, color=HEADER_FG),
                   fill=header_fill(),
                   alignment=center_align(wrap=True),
                   border=thin_border())
    ws.row_dimensions[hdr_row].height = 28

    # Data rows
    data_start = hdr_row + 1
    gpr = model.gross_potential_rent()

    for row_idx, unit in enumerate(model.units):
        r = data_start + row_idx
        is_vacant = unit.is_vacant
        row_fill = PatternFill("solid", fgColor="FFF2CC") if is_vacant else None

        # Annual rent formula (col 7) references col 6 (monthly rent)
        monthly_col = get_column_letter(6)   # F
        annual_col  = get_column_letter(7)   # G
        rsf_col     = get_column_letter(8)   # H  rent/sf
        mo_remaining_col = get_column_letter(13)  # M

        data = [
            (2,  unit.unit_id,         input_font(),   None,              left_align(indent=1)),
            (3,  unit.tenant or "Vacant", input_font(), None,             left_align(indent=1)),
            (4,  unit.unit_type,       input_font(),   None,              center_align()),
            (5,  unit.sf,              input_font(),   "#,##0",           right_align()),
            (6,  unit.monthly_rent,    input_font(),   "$#,##0",          right_align()),
            # Annual rent = monthly × 12 (formula)
            (7,  f"={monthly_col}{r}*12",  formula_font(), "$#,##0",     right_align()),
            # Rent per sf = annual / sf (formula)
            (8,  f"={annual_col}{r}/{get_column_letter(5)}{r}" if unit.sf > 0 else 0,
                 formula_font(), "$#,##0.00",  right_align()),
            (9,  unit.lease_start,     input_font(),   "MM/DD/YYYY",      center_align()),
            (10, unit.lease_end,       input_font(),   "MM/DD/YYYY",      center_align()),
            (11, unit.annual_escalation, input_font(), "0.0%",           center_align()),
            (12, unit.status,          input_font(),   None,              center_align()),
        ]
        for col_offset, val, fnt, fmt, aln in data:
            cell = ws.cell(row=r, column=col_offset, value=val)
            cell.font = fnt
            if fmt:
                cell.number_format = fmt
            cell.alignment = aln
            if row_fill and not is_vacant:
                pass
            elif is_vacant:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.border = thin_border()

        # Months remaining formula (col 13) — Excel formula using lease_end (col J = 10)
        lease_end_ref = f"{get_column_letter(10)}{r}"
        ws.cell(row=r, column=13,
                value=f"=MAX(0,({lease_end_ref}-TODAY())/30.44)").number_format = "#,##0.0"
        ws.cell(row=r, column=13).font = formula_font()
        ws.cell(row=r, column=13).alignment = right_align()
        ws.cell(row=r, column=13).border = thin_border()

        # % of GPR formula (col 14) — annual rent / sum of all annual rents
        ann_cell = f"{annual_col}{r}"
        ws.cell(row=r, column=14,
                value=f"={ann_cell}/SUM({annual_col}{data_start}:{annual_col}{data_start + len(model.units) - 1})").number_format = "0.0%"
        ws.cell(row=r, column=14).font = formula_font()
        ws.cell(row=r, column=14).alignment = right_align()
        ws.cell(row=r, column=14).border = thin_border()

        ws.row_dimensions[r].height = 16

    # Totals row
    total_row = data_start + len(model.units)
    ws.row_dimensions[total_row].height = 18
    write_cell(ws, total_row, 2, "TOTAL / WEIGHTED AVG",
               font=Font(name="Calibri", bold=True, size=9),
               fill=total_fill(), alignment=left_align(indent=1), border=thin_border())

    # SF total
    write_cell(ws, total_row, 5,
               f"=SUM({get_column_letter(5)}{data_start}:{get_column_letter(5)}{total_row - 1})",
               font=Font(name="Calibri", bold=True, size=9), fill=total_fill(),
               alignment=right_align(), border=thin_border()).number_format = "#,##0"

    # Monthly rent total
    write_cell(ws, total_row, 6,
               f"=SUM({get_column_letter(6)}{data_start}:{get_column_letter(6)}{total_row - 1})",
               font=Font(name="Calibri", bold=True, size=9), fill=total_fill(),
               alignment=right_align(), border=thin_border()).number_format = "$#,##0"

    # Annual rent total
    write_cell(ws, total_row, 7,
               f"=SUM({get_column_letter(7)}{data_start}:{get_column_letter(7)}{total_row - 1})",
               font=Font(name="Calibri", bold=True, size=9), fill=total_fill(),
               alignment=right_align(), border=thin_border()).number_format = "$#,##0"

    # Weighted avg rent/sf = total annual rent / total SF
    write_cell(ws, total_row, 8,
               f"={get_column_letter(7)}{total_row}/{get_column_letter(5)}{total_row}",
               font=Font(name="Calibri", bold=True, size=9), fill=total_fill(),
               alignment=right_align(), border=thin_border()).number_format = "$#,##0.00"

    # % of GPR total — should be 100%
    write_cell(ws, total_row, 14,
               f"=SUM({get_column_letter(14)}{data_start}:{get_column_letter(14)}{total_row - 1})",
               font=Font(name="Calibri", bold=True, size=9), fill=total_fill(),
               alignment=right_align(), border=thin_border()).number_format = "0.0%"

    # Empty remaining totals
    for c in [3, 4, 9, 10, 11, 12, 13]:
        ws.cell(row=total_row, column=c).fill = total_fill()
        ws.cell(row=total_row, column=c).border = thin_border()

    # Freeze panes
    ws.freeze_panes = f"B{hdr_row + 1}"

    # Column widths
    col_widths = [3, 10, 24, 10, 9, 14, 14, 12, 12, 12, 10, 14, 15, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Footer
    footer_row = total_row + 2
    ws.merge_cells(f"B{footer_row}:N{footer_row}")
    write_cell(ws, footer_row, 2,
               "Blue cells are hardcoded inputs. Black cells contain formulas. Vacant units highlighted in yellow.",
               font=Font(name="Calibri", size=8, italic=True, color="888888"),
               alignment=left_align())


# ---------------------------------------------------------------------------
# Sheet 3 — Pro Forma
# ---------------------------------------------------------------------------

def build_pro_forma_sheet(ws, model: CREDealModel):
    """
    5-Year Pro Forma using Excel formulas referencing assumption cells.

    All inputs (growth rates, base values) are placed as hardcoded blue inputs
    in a dedicated assumptions block; projection columns use formula references.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8

    # Title
    ws.merge_cells("B2:H2")
    write_cell(ws, 2, 2, "5-YEAR PRO FORMA INCOME STATEMENT",
               font=Font(name="Calibri", bold=True, size=14, color=HEADER_FG),
               fill=header_fill(), alignment=center_align())
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:H3")
    write_cell(ws, 3, 2, "123 Delancey Street  |  NYC Mixed-Use  |  Projection Period: 2026–2030",
               font=Font(name="Calibri", size=9, italic=True),
               fill=PatternFill("solid", fgColor=SECTION_BG),
               alignment=center_align())
    ws.row_dimensions[3].height = 14

    # ---------- Assumption Cells (blue inputs) ----------
    # Row 5: Labels; Row 6: Values
    section_header(ws, 5, 2, "KEY ASSUMPTIONS (Blue = Input)", span_end_col=8)
    ws.row_dimensions[5].height = 18

    ASSUMPTIONS = {
        "base_gpr":         (7, 2,  model.gross_potential_rent(),    "$#,##0",  "Base Gross Potential Rent ($)"),
        "vac_rate":         (7, 4,  0.05,                            "0.0%",    "Vacancy Rate"),
        "credit_loss":      (7, 6,  0.01,                            "0.0%",    "Credit Loss Rate"),
        "rent_growth":      (9, 2,  0.03,                            "0.0%",    "Annual Rent Growth"),
        "opex_growth":      (9, 4,  0.03,                            "0.0%",    "Annual OpEx Growth"),
        "mgmt_fee_pct":     (9, 6,  0.04,                            "0.0%",    "Mgmt Fee (% of EGI)"),
        "re_taxes":         (11, 2, model.opex.real_estate_taxes,    "$#,##0",  "Real Estate Taxes ($)"),
        "insurance":        (11, 4, model.opex.insurance,            "$#,##0",  "Insurance ($)"),
        "repairs":          (11, 6, model.opex.repairs_maintenance,  "$#,##0",  "Repairs & Maint. ($)"),
        "utilities":        (13, 2, model.opex.utilities,            "$#,##0",  "Utilities ($)"),
        "gna":              (13, 4, model.opex.general_admin,        "$#,##0",  "G&A ($)"),
        "reserves":         (13, 6, model.opex.reserves,             "$#,##0",  "Reserves ($)"),
    }

    # Named cell addresses (for formula references)
    ASSUMP_ADDR = {}
    for key, (arow, acol, val, fmt, label) in ASSUMPTIONS.items():
        label_cell = ws.cell(row=arow - 1, column=acol, value=label)
        label_cell.font = Font(name="Calibri", size=9, bold=False, color="444444")
        label_cell.alignment = left_align()

        val_cell = ws.cell(row=arow, column=acol, value=val)
        val_cell.font = input_font()
        val_cell.number_format = fmt
        val_cell.alignment = right_align()
        val_cell.fill = PatternFill("solid", fgColor="EBF3FB")
        ASSUMP_ADDR[key] = f"{get_column_letter(acol)}{arow}"
        ws.row_dimensions[arow].height = 15
        ws.row_dimensions[arow - 1].height = 14

    # ---------- Column Headers (Year 1–5) ----------
    HDR_ROW = 16
    section_header(ws, HDR_ROW - 1, 2, "PROJECTED INCOME STATEMENT", span_end_col=8)
    ws.row_dimensions[HDR_ROW - 1].height = 18

    write_cell(ws, HDR_ROW, 2, "Line Item",
               font=Font(name="Calibri", bold=True, size=9, color=HEADER_FG),
               fill=header_fill(), alignment=left_align(indent=1))
    base_year = model.analysis_date.year
    for i, yr in enumerate(range(base_year + 1, base_year + 6)):
        col = 3 + i
        write_cell(ws, HDR_ROW, col, f"Year {i + 1} ({yr})",
                   font=Font(name="Calibri", bold=True, size=9, color=HEADER_FG),
                   fill=header_fill(), alignment=center_align())
    ws.row_dimensions[HDR_ROW].height = 20

    # ---------- Income Statement Rows ----------
    # Each year uses compounded-growth formulas referencing assumption cells.
    # Y1 formulas shown; Y2-Y5 just increment the exponent.

    def yr_formula(base_addr, growth_addr, year_num):
        """Compound growth formula: base * (1+g)^n"""
        return f"={base_addr}*(1+{growth_addr})^{year_num}"

    def vcl_formula(gpr_cell, vac_addr, cl_addr):
        return f"=({vac_addr}+{cl_addr})*{gpr_cell}"

    DATA_START = HDR_ROW + 1
    row = DATA_START
    col_letters = {yr: get_column_letter(3 + i) for i, yr in enumerate(range(1, 6))}

    # ---- GPR ----
    ws.cell(row=row, column=2, value="Gross Potential Rent (GPR)").font = formula_font(bold=True)
    ws.cell(row=row, column=2).alignment = left_align(indent=1)
    gpr_cells = {}
    for yr in range(1, 6):
        col = 3 + yr - 1
        cell = ws.cell(
            row=row, column=col,
            value=yr_formula(ASSUMP_ADDR["base_gpr"], ASSUMP_ADDR["rent_growth"], yr),
        )
        cell.font = formula_font(bold=True)
        cell.number_format = "$#,##0"
        cell.alignment = right_align()
        gpr_cells[yr] = f"{col_letters[yr]}{row}"
    ws.row_dimensions[row].height = 16
    gpr_row = row
    row += 1

    # ---- Vacancy & Credit Loss ----
    ws.cell(row=row, column=2, value="Less: Vacancy & Credit Loss").font = formula_font()
    ws.cell(row=row, column=2).alignment = left_align(indent=2)
    vcl_cells = {}
    for yr in range(1, 6):
        col = 3 + yr - 1
        cell = ws.cell(
            row=row, column=col,
            value=vcl_formula(gpr_cells[yr], ASSUMP_ADDR["vac_rate"], ASSUMP_ADDR["credit_loss"]),
        )
        cell.font = formula_font()
        cell.number_format = "$#,##0"
        cell.alignment = right_align()
        vcl_cells[yr] = f"{col_letters[yr]}{row}"
    ws.row_dimensions[row].height = 15
    vcl_row = row
    row += 1

    # ---- EGI ----
    ws.cell(row=row, column=2, value="Effective Gross Income (EGI)").font = formula_font(bold=True)
    ws.cell(row=row, column=2).alignment = left_align(indent=1)
    ws.cell(row=row, column=2).fill = section_fill()
    egi_cells = {}
    for yr in range(1, 6):
        col = 3 + yr - 1
        gpr_ref = gpr_cells[yr]
        vcl_ref = vcl_cells[yr]
        cell = ws.cell(row=row, column=col, value=f"={gpr_ref}-{vcl_ref}")
        cell.font = formula_font(bold=True)
        cell.number_format = "$#,##0"
        cell.alignment = right_align()
        cell.fill = section_fill()
        egi_cells[yr] = f"{col_letters[yr]}{row}"
    ws.row_dimensions[row].height = 16
    egi_row = row
    row += 2  # blank spacer

    # ---- Operating Expenses ----
    section_header(ws, row, 2, "Operating Expenses", span_end_col=8)
    ws.row_dimensions[row].height = 16
    row += 1

    expense_keys = [
        ("Real Estate Taxes",       ASSUMP_ADDR["re_taxes"],   ASSUMP_ADDR["opex_growth"]),
        ("Insurance",               ASSUMP_ADDR["insurance"],  ASSUMP_ADDR["opex_growth"]),
        ("Repairs & Maintenance",   ASSUMP_ADDR["repairs"],    ASSUMP_ADDR["opex_growth"]),
        ("Utilities",               ASSUMP_ADDR["utilities"],  ASSUMP_ADDR["opex_growth"]),
        ("General & Administrative",ASSUMP_ADDR["gna"],        ASSUMP_ADDR["opex_growth"]),
        ("Reserves",                ASSUMP_ADDR["reserves"],   ASSUMP_ADDR["opex_growth"]),
    ]

    opex_row_refs = []  # list of {yr: cell_ref} per expense line
    for label, base_addr, growth_addr in expense_keys:
        ws.cell(row=row, column=2, value=label).font = formula_font(size=9)
        ws.cell(row=row, column=2).alignment = left_align(indent=2)
        line_cells = {}
        for yr in range(1, 6):
            col = 3 + yr - 1
            cell = ws.cell(
                row=row, column=col,
                value=yr_formula(base_addr, growth_addr, yr),
            )
            cell.font = formula_font(size=9)
            cell.number_format = "$#,##0"
            cell.alignment = right_align()
            line_cells[yr] = f"{col_letters[yr]}{row}"
        opex_row_refs.append(line_cells)
        ws.row_dimensions[row].height = 15
        row += 1

    # ---- Management Fee ----
    ws.cell(row=row, column=2, value="Management Fee").font = formula_font(size=9)
    ws.cell(row=row, column=2).alignment = left_align(indent=2)
    mgmt_cells = {}
    for yr in range(1, 6):
        col = 3 + yr - 1
        egi_ref = egi_cells[yr]
        cell = ws.cell(row=row, column=col, value=f"={ASSUMP_ADDR['mgmt_fee_pct']}*{egi_ref}")
        cell.font = formula_font(size=9)
        cell.number_format = "$#,##0"
        cell.alignment = right_align()
        mgmt_cells[yr] = f"{col_letters[yr]}{row}"
    opex_row_refs.append(mgmt_cells)
    ws.row_dimensions[row].height = 15
    row += 1

    # ---- Total OpEx ----
    ws.cell(row=row, column=2, value="Total Operating Expenses").font = formula_font(bold=True)
    ws.cell(row=row, column=2).alignment = left_align(indent=1)
    ws.cell(row=row, column=2).fill = section_fill()
    total_opex_cells = {}
    for yr in range(1, 6):
        col = 3 + yr - 1
        yr_refs = "+".join(line_dict[yr] for line_dict in opex_row_refs)
        cell = ws.cell(row=row, column=col, value=f"={yr_refs}")
        cell.font = formula_font(bold=True)
        cell.number_format = "$#,##0"
        cell.alignment = right_align()
        cell.fill = section_fill()
        total_opex_cells[yr] = f"{col_letters[yr]}{row}"
    ws.row_dimensions[row].height = 16
    row += 2

    # ---- NOI ----
    ws.cell(row=row, column=2, value="NET OPERATING INCOME (NOI)").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=2).alignment = left_align(indent=1)
    ws.cell(row=row, column=2).fill = total_fill()
    noi_cells = {}
    for yr in range(1, 6):
        col = 3 + yr - 1
        egi_ref = egi_cells[yr]
        opex_ref = total_opex_cells[yr]
        cell = ws.cell(row=row, column=col, value=f"={egi_ref}-{opex_ref}")
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.number_format = "$#,##0"
        cell.alignment = right_align()
        cell.fill = total_fill()
        noi_cells[yr] = f"{col_letters[yr]}{row}"
        cell.border = thick_bottom_border()
    ws.row_dimensions[row].height = 20
    noi_row = row
    row += 2

    # ---- Expense Ratio ----
    ws.cell(row=row, column=2, value="Expense Ratio (OpEx / EGI)").font = formula_font(size=9)
    ws.cell(row=row, column=2).alignment = left_align(indent=2)
    for yr in range(1, 6):
        col = 3 + yr - 1
        cell = ws.cell(row=row, column=col,
                       value=f"={total_opex_cells[yr]}/{egi_cells[yr]}")
        cell.font = formula_font(size=9)
        cell.number_format = "0.0%"
        cell.alignment = right_align()
    ws.row_dimensions[row].height = 15
    row += 1

    # ---- NOI / SF ----
    total_sf = model.total_sf
    ws.cell(row=row, column=2, value="NOI / SF").font = formula_font(size=9)
    ws.cell(row=row, column=2).alignment = left_align(indent=2)
    for yr in range(1, 6):
        col = 3 + yr - 1
        cell = ws.cell(row=row, column=col, value=f"={noi_cells[yr]}/{total_sf}")
        cell.font = formula_font(size=9)
        cell.number_format = "$#,##0.00"
        cell.alignment = right_align()
    ws.row_dimensions[row].height = 15

    # Column widths
    set_col_width(ws, 2, 32)
    for c in range(3, 8):
        set_col_width(ws, c, 18)

    ws.freeze_panes = f"C{HDR_ROW + 1}"

    # Footer
    ws.merge_cells(f"B{row + 2}:H{row + 2}")
    write_cell(ws, row + 2, 2,
               "Formulas reference assumption cells above. Modify blue input cells to run scenarios.",
               font=Font(name="Calibri", size=8, italic=True, color="888888"),
               alignment=left_align())


# ---------------------------------------------------------------------------
# Sheet 4 — Debt Sizing
# ---------------------------------------------------------------------------

def build_debt_sizing_sheet(ws, model: CREDealModel):
    """3-lender comparison table with recommendation."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8

    noi = model.net_operating_income()
    value = model.direct_cap_value(0.055)
    engine = DebtSizingEngine(noi=noi, property_value=value)
    table = engine.comparison_table()
    rec = engine.recommend()

    # Title
    ws.merge_cells("B2:K2")
    write_cell(ws, 2, 2, "DEBT SIZING — LENDER COMPARISON",
               font=Font(name="Calibri", bold=True, size=14, color=HEADER_FG),
               fill=header_fill(), alignment=center_align())
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:K3")
    write_cell(ws, 3, 2,
               f"NOI: ${noi:,.0f}  |  Property Value: ${value:,.0f}  |  Going-In Cap Rate: 5.50%",
               font=Font(name="Calibri", size=9, italic=True),
               fill=PatternFill("solid", fgColor=SECTION_BG),
               alignment=center_align())
    ws.row_dimensions[3].height = 14

    # -- Lender Input Assumptions (rows 5–13) --
    section_header(ws, 5, 2, "LENDER PARAMETERS (Inputs)", span_end_col=11)
    ws.row_dimensions[5].height = 18

    param_headers = ["Parameter", "Bank", "Debt Fund", "Agency (Fannie/Freddie)"]
    for col_idx, h in enumerate(param_headers, start=2):
        write_cell(ws, 6, col_idx, h,
                   font=Font(name="Calibri", bold=True, size=9, color=HEADER_FG),
                   fill=header_fill(SUBHEADER_BG if col_idx == 2 else HEADER_BG),
                   alignment=center_align())
    ws.row_dimensions[6].height = 18

    from models.debt_sizing_engine import BANK_LENDER, DEBT_FUND_LENDER, AGENCY_LENDER
    lender_profiles = [BANK_LENDER, DEBT_FUND_LENDER, AGENCY_LENDER]

    params = [
        ("Minimum DSCR",      [f"{l.min_dscr:.2f}x" for l in lender_profiles],      "0.00"),
        ("Maximum LTV",       [f"{l.max_ltv:.0%}" for l in lender_profiles],         "0%"),
        ("Min. Debt Yield",   [f"{l.min_debt_yield:.1%}" for l in lender_profiles],  "0.0%"),
        ("Interest Rate",     [f"{l.interest_rate:.2%}" for l in lender_profiles],   "0.00%"),
        ("Amortization",      ["25yr Amort", "Interest Only", "30yr Amort"],          None),
        ("Loan Term",         [f"{l.loan_term_years}yr" for l in lender_profiles],   None),
    ]

    prow = 7
    for label, values, _ in params:
        write_cell(ws, prow, 2, label, font=formula_font(size=9), alignment=left_align())
        for i, v in enumerate(values):
            write_cell(ws, prow, 3 + i, v, font=input_font(size=9), alignment=center_align(),
                       fill=PatternFill("solid", fgColor="EBF3FB"))
        ws.row_dimensions[prow].height = 15
        prow += 1

    # -- Results Table --
    section_header(ws, prow + 1, 2, "SIZING RESULTS (Formulas)", span_end_col=11)
    ws.row_dimensions[prow + 1].height = 18

    result_headers = [
        "Metric", "Bank", "Debt Fund", "Agency", "Recommended"
    ]
    rhdr = prow + 2
    for col_idx, h in enumerate(result_headers, start=2):
        bg = HEADER_BG if col_idx > 2 else SUBHEADER_BG
        write_cell(ws, rhdr, col_idx, h,
                   font=Font(name="Calibri", bold=True, size=9, color=HEADER_FG),
                   fill=header_fill(bg),
                   alignment=center_align())
    ws.row_dimensions[rhdr].height = 18

    result_data = [
        ("Max Loan ($)",         [f"${r['Max Loan ($)']:,.0f}" for r in table],      "$#,##0"),
        ("Annual Debt Svc ($)",  [f"${r['Annual Debt Service ($)']:,.0f}" for r in table], "$#,##0"),
        ("Actual DSCR",         [f"{r['Actual DSCR']:.2f}x" for r in table],        None),
        ("Actual LTV",           [f"{r['Actual LTV']:.1%}" for r in table],         "0.0%"),
        ("Actual Debt Yield",    [f"{r['Actual Debt Yield']:.1%}" for r in table],  "0.0%"),
        ("Binding Constraint",   [r['Binding Constraint'] for r in table],          None),
    ]

    drow = rhdr + 1
    for label, values, _ in result_data:
        is_loan = label == "Max Loan ($)"
        fnt = Font(name="Calibri", bold=is_loan, size=9)
        fill = total_fill() if is_loan else None
        write_cell(ws, drow, 2, label, font=fnt, fill=fill, alignment=left_align())

        # Sort table by lender name to match column order: Bank, Debt Fund, Agency
        lender_order = ["Bank", "Debt Fund", "Agency (Fannie/Freddie)"]
        sorted_table = sorted(table, key=lambda r: lender_order.index(r["Lender"])
                              if r["Lender"] in lender_order else 99)

        for i, row_data in enumerate(sorted_table):
            val = {
                "Max Loan ($)":        row_data["Max Loan ($)"],
                "Annual Debt Svc ($)": row_data["Annual Debt Service ($)"],
                "Actual DSCR":         row_data["Actual DSCR"],
                "Actual LTV":          row_data["Actual LTV"],
                "Actual Debt Yield":   row_data["Actual Debt Yield"],
                "Binding Constraint":  row_data["Binding Constraint"],
            }.get(label, "")

            cell = ws.cell(row=drow, column=3 + i, value=val)
            cell.font = fnt
            if fill:
                cell.fill = fill
            if isinstance(val, float):
                if "DSCR" in label:
                    cell.number_format = "0.00"
                elif "LTV" in label or "Yield" in label:
                    cell.number_format = "0.0%"
                elif "$" in label:
                    cell.number_format = "$#,##0"
            cell.alignment = center_align()

        # Recommended column (col 6)
        rec_val = {
            "Max Loan ($)":        rec["max_loan"],
            "Annual Debt Svc ($)": rec["annual_debt_service"],
            "Actual DSCR":         rec["actual_dscr"],
            "Actual LTV":          rec["actual_ltv"],
            "Actual Debt Yield":   rec["actual_debt_yield"],
            "Binding Constraint":  rec["binding_constraint"],
        }.get(label, "")

        rec_cell = ws.cell(row=drow, column=6, value=rec_val)
        rec_cell.font = Font(name="Calibri", bold=True, size=9, color="008000")
        if isinstance(rec_val, float):
            if "DSCR" in label:
                rec_cell.number_format = "0.00"
            elif "LTV" in label or "Yield" in label:
                rec_cell.number_format = "0.0%"
            elif "$" in label:
                rec_cell.number_format = "$#,##0"
        rec_cell.alignment = center_align()
        if is_loan:
            rec_cell.fill = total_fill()

        ws.row_dimensions[drow].height = 16
        drow += 1

    # Recommendation box
    drow += 1
    section_header(ws, drow, 2, "RECOMMENDATION", span_end_col=11)
    ws.row_dimensions[drow].height = 18
    drow += 1

    ws.merge_cells(f"B{drow}:K{drow}")
    write_cell(ws, drow, 2, rec.get("rationale", "See recommendation above."),
               font=Font(name="Calibri", size=10, bold=True, color="008000"),
               fill=PatternFill("solid", fgColor="E2EFDA"),
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1))
    ws.row_dimensions[drow].height = 36

    # Column widths
    set_col_width(ws, 2, 28)
    for c in [3, 4, 5, 6]:
        set_col_width(ws, c, 22)
    for c in range(7, 12):
        set_col_width(ws, c, 14)

    ws.freeze_panes = f"B{rhdr + 1}"


# ---------------------------------------------------------------------------
# Sheet 5 — Sensitivity
# ---------------------------------------------------------------------------

def build_sensitivity_sheet(ws, sa: SensitivityAnalysis):
    """Sensitivity tables with conditional formatting."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 8

    # Title
    ws.merge_cells("B2:Q2")
    write_cell(ws, 2, 2, "SENSITIVITY ANALYSIS",
               font=Font(name="Calibri", bold=True, size=14, color=HEADER_FG),
               fill=header_fill(), alignment=center_align())
    ws.row_dimensions[2].height = 28

    base = sa.base_metrics()
    ws.merge_cells("B3:Q3")
    write_cell(ws, 3, 2,
               f"Base NOI: ${base['noi']:,.0f}  |  Base Value: ${base['property_value']:,.0f}  |  DSCR: {base['actual_dscr']:.2f}x  |  LTV: {base['actual_ltv']:.1%}",
               font=Font(name="Calibri", size=9, italic=True),
               fill=PatternFill("solid", fgColor=SECTION_BG),
               alignment=center_align())

    current_row = 5

    def write_one_way_table(results, title, variable_header):
        nonlocal current_row

        section_header(ws, current_row, 2, title, span_end_col=10)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        cols = [variable_header, "NOI ($)", "Property Value ($)",
                "Max Loan ($)", "DSCR", "LTV", "Debt Yield", "Risk Flag"]
        for ci, h in enumerate(cols, start=2):
            write_cell(ws, current_row, ci, h,
                       font=Font(name="Calibri", bold=True, size=9, color=HEADER_FG),
                       fill=header_fill(SUBHEADER_BG),
                       alignment=center_align(wrap=True))
        ws.row_dimensions[current_row].height = 28
        data_start = current_row + 1
        current_row += 1

        for r in results:
            flag_str = ""
            if r.dscr_flag == "BREACH" or r.ltv_flag == "BREACH":
                flag_str = "BREACH"
            elif r.dscr_flag == "WARNING" or r.ltv_flag == "WARNING":
                flag_str = "WARNING"

            row_fill = None
            if flag_str == "BREACH":
                row_fill = PatternFill("solid", fgColor="FFCCCC")
            elif flag_str == "WARNING":
                row_fill = PatternFill("solid", fgColor="FFFACD")

            cells = [
                (r.variable_label, None, left_align()),
                (r.noi, "$#,##0", right_align()),
                (r.property_value, "$#,##0", right_align()),
                (r.max_loan, "$#,##0", right_align()),
                (r.actual_dscr, "0.00", right_align()),
                (r.actual_ltv, "0.0%", right_align()),
                (r.actual_debt_yield, "0.0%", right_align()),
                (flag_str or "OK", None, center_align()),
            ]
            for ci, (val, fmt, aln) in enumerate(cells, start=2):
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.font = input_font() if ci == 2 else formula_font(size=9)
                if fmt:
                    cell.number_format = fmt
                cell.alignment = aln
                if row_fill:
                    cell.fill = row_fill

            ws.row_dimensions[current_row].height = 15
            current_row += 1

        # Conditional formatting — DSCR column (col F = 7): red if < 1.0
        dscr_col = get_column_letter(7)
        dscr_range = f"{dscr_col}{data_start}:{dscr_col}{current_row - 1}"
        ws.conditional_formatting.add(
            dscr_range,
            CellIsRule(operator="lessThan", formula=["1.0"],
                       fill=PatternFill("solid", fgColor="FFCCCC"))
        )
        ws.conditional_formatting.add(
            dscr_range,
            CellIsRule(operator="between", formula=["1.0", "1.25"],
                       fill=PatternFill("solid", fgColor="FFFACD"))
        )

        current_row += 2

    # One-way tables
    write_one_way_table(sa.rent_growth_sensitivity(), "1. RENT GROWTH RATE SENSITIVITY", "Rent Growth Shift")
    write_one_way_table(sa.vacancy_sensitivity(), "2. VACANCY RATE SENSITIVITY", "Vacancy Rate")
    write_one_way_table(sa.interest_rate_sensitivity(), "3. INTEREST RATE SENSITIVITY", "Rate Shift")
    write_one_way_table(sa.cap_rate_sensitivity(), "4. CAP RATE SENSITIVITY", "Cap Rate Shift")
    write_one_way_table(sa.opex_growth_sensitivity(), "5. OPEX GROWTH SENSITIVITY", "OpEx Growth Shift")

    # Two-way table: rent growth vs vacancy (NOI)
    def write_two_way(two_way_data, title, fmt="$#,##0"):
        nonlocal current_row
        section_header(ws, current_row, 2, title, span_end_col=15)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Axis labels
        ws.cell(row=current_row, column=2,
                value=f"↓ {two_way_data['row_axis']}  /  → {two_way_data['col_axis']}").font = \
            Font(name="Calibri", size=8, italic=True)
        current_row += 1

        # Column headers
        write_cell(ws, current_row, 2, "",
                   fill=PatternFill("solid", fgColor=SUBHEADER_BG))
        for ci, label in enumerate(two_way_data["col_labels"], start=3):
            write_cell(ws, current_row, ci, label,
                       font=Font(name="Calibri", bold=True, size=8),
                       fill=PatternFill("solid", fgColor=SUBHEADER_BG),
                       alignment=center_align())
        ws.row_dimensions[current_row].height = 20
        table_start_row = current_row + 1
        current_row += 1

        for row_label, row_vals, row_flags in zip(
            two_way_data["row_labels"], two_way_data["table"], two_way_data["flags"]
        ):
            write_cell(ws, current_row, 2, row_label,
                       font=Font(name="Calibri", bold=True, size=8),
                       fill=PatternFill("solid", fgColor=SUBHEADER_BG),
                       alignment=center_align())
            for ci, (val, flag) in enumerate(zip(row_vals, row_flags), start=3):
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.font = formula_font(size=8)
                if fmt:
                    cell.number_format = fmt
                cell.alignment = right_align()
                if flag == "BREACH":
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                elif flag == "WARNING":
                    cell.fill = PatternFill("solid", fgColor="FFFACD")
            ws.row_dimensions[current_row].height = 15
            current_row += 1

        # Color scale conditional formatting on the entire data range
        n_cols = len(two_way_data["col_labels"])
        n_rows = len(two_way_data["row_labels"])
        data_range = (
            f"{get_column_letter(3)}{table_start_row}:"
            f"{get_column_letter(3 + n_cols - 1)}{table_start_row + n_rows - 1}"
        )
        inverted = two_way_data["metric"] in ("actual_ltv",)
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min", start_color="F8696B" if not inverted else "63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B" if not inverted else "F8696B",
            )
        )
        current_row += 2

    tw_noi = sa.two_way_rent_vacancy(metric="noi")
    write_two_way(tw_noi, "6. 2D: RENT GROWTH × VACANCY — NOI ($)", fmt="$#,##0")

    tw_dscr = sa.two_way_rent_vacancy(metric="actual_dscr")
    write_two_way(tw_dscr, "7. 2D: RENT GROWTH × VACANCY — DSCR (x)", fmt="0.00")

    tw_loan = sa.two_way_rate_caprate(metric="max_loan")
    write_two_way(tw_loan, "8. 2D: INTEREST RATE × CAP RATE — MAX LOAN ($)", fmt="$#,##0")

    # Column widths
    set_col_width(ws, 2, 22)
    for c in range(3, 16):
        set_col_width(ws, c, 14)

    ws.freeze_panes = "B5"


# ---------------------------------------------------------------------------
# Main Entry Point
# ---------------------------------------------------------------------------

def main():
    # Ensure output directory exists
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "sample_deal_model.xlsx")

    print("Building deal model...")
    model = build_deal_model()

    print("Running sensitivity analysis...")
    sa = SensitivityAnalysis(
        deal_model=model,
        base_cap_rate=0.055,
        base_interest_rate=0.0575,
        base_opex_growth_rate=0.03,
    )
    base = sa.base_metrics()

    print("Creating workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    # Create sheets
    ws_summary  = wb.create_sheet("Summary")
    ws_rent     = wb.create_sheet("Rent Roll")
    ws_pf       = wb.create_sheet("Pro Forma")
    ws_debt     = wb.create_sheet("Debt Sizing")
    ws_sens     = wb.create_sheet("Sensitivity")

    print("  → Building Summary sheet...")
    build_summary_sheet(ws_summary, model, sa, base)

    print("  → Building Rent Roll sheet...")
    build_rent_roll_sheet(ws_rent, model)

    print("  → Building Pro Forma sheet...")
    build_pro_forma_sheet(ws_pf, model)

    print("  → Building Debt Sizing sheet...")
    build_debt_sizing_sheet(ws_debt, model)

    print("  → Building Sensitivity sheet...")
    build_sensitivity_sheet(ws_sens, sa)

    print(f"Saving workbook to {output_path}...")
    wb.save(output_path)
    print(f"\nSuccess! Excel model saved to:\n  {output_path}\n")
    print("Summary of what was built:")
    print(f"  NOI              : ${model.net_operating_income():,.0f}")
    print(f"  Property Value   : ${model.direct_cap_value(0.055):,.0f}")
    print(f"  Total SF         : {model.total_sf:,.0f}")
    print(f"  Physical Occ.    : {model.physical_occupancy:.1%}")


if __name__ == "__main__":
    main()
